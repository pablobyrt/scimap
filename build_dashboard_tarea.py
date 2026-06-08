"""
Dashboard bibliométrico — U. Chile IA 2022-2025
"""
import pandas as pd, re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ── rutas ─────────────────────────────────────────────────────────────────────
SRC  = r"C:\Users\Pablo\Desktop\Magister\Segundo Semestre\Informetría\Trabajo 2\Trabajo 2 - con sexo estimado.xlsx"
DEST = r"C:\Users\Pablo\Desktop\Magister\Segundo Semestre\Informetría\Trabajo 2\Dashboard Bibliométrico - U. Chile IA 2022-2025.xlsx"

# ── estilos ───────────────────────────────────────────────────────────────────
C_DARK  = "1E3A5F"
C_BLUE  = "2563EB"
C_LBLUE = "DBEAFE"
C_MINT  = "D1FAE5"
C_AMBE  = "FEF3C7"
C_PINK  = "FCE7F3"
C_WHITE = "FFFFFF"
C_GRAY  = "F8FAFC"
C_DGRAY = "64748B"

def fill(hex_):       return PatternFill("solid", start_color=hex_)
def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_bottom():
    s = Side(style="thin", color="E2E8F0")
    return Border(bottom=s)
def full_border():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row, cols, bg=C_DARK, fg=C_WHITE, size=11):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = font(bold=True, size=size, color=fg)
        cell.alignment = align("center")

def style_data_row(ws, row, cols, bg=C_WHITE, stripe=C_GRAY, is_even=True):
    bg_ = stripe if is_even else bg
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg_)
        cell.font = font(size=10)
        cell.alignment = align("center")
        cell.border = border_bottom()

def kpi_box(ws, row, col, value, label, bg, fg="FFFFFF"):
    """Caja KPI de 2 cols x 4 rows."""
    for r in range(row, row + 4):
        for c in range(col, col + 2):
            ws.cell(r, c).fill = fill(bg)
    ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col+1)
    v = ws.cell(row, col)
    v.value = value
    v.font = Font(name="Calibri", bold=True, size=22, color=fg)
    v.alignment = align("center")
    v.fill = fill(bg)
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+3, end_column=col+1)
    l = ws.cell(row+2, col)
    l.value = label
    l.font = Font(name="Calibri", size=9, color=fg, italic=True)
    l.alignment = align("center")
    l.fill = fill(bg)

def section_title(ws, row, col, title):
    cell = ws.cell(row, col)
    cell.value = title
    cell.font = Font(name="Calibri", bold=True, size=13, color=C_DARK)
    cell.alignment = align("left")
    cell.fill = fill(C_WHITE)

# ── carga y procesa datos ─────────────────────────────────────────────────────
df = pd.read_excel(SRC, sheet_name=0)
df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
df["Cited by"] = pd.to_numeric(df["Cited by"], errors="coerce").fillna(0)

COUNTRIES = [
    "Chile","United States","Spain","Brazil","France","Germany","United Kingdom",
    "Argentina","Colombia","Mexico","Peru","Italy","Australia","Canada","China",
    "Japan","South Korea","Netherlands","Sweden","Switzerland","Portugal","Uruguay",
    "Ecuador","India","Denmark","Norway","Finland","Belgium","Austria","Poland",
    "Israel","Singapore","Taiwan","Malaysia","Indonesia","New Zealand","Russia",
]

def extract_countries(aff):
    if not isinstance(aff, str): return []
    found = []
    for entry in aff.split(";"):
        parts = [p.strip() for p in entry.split(",")]
        for p in reversed(parts):
            if p in COUNTRIES:
                found.append(p)
                break
    return list(dict.fromkeys(found))

df["countries"] = df["Affiliations"].apply(extract_countries)

# Métricas base
total_docs   = len(df)
total_cites  = int(df["Cited by"].sum())
avg_cites    = round(df["Cited by"].mean(), 1)
cites_sorted = sorted(df["Cited by"], reverse=True)
h_index      = sum(1 for i, c in enumerate(cites_sorted, 1) if c >= i)
intl_collabs = df[df["countries"].apply(lambda c: any(x != "Chile" for x in c))]
pct_intl     = round(100 * len(intl_collabs) / total_docs, 1)

# Por año
by_year = df.groupby("Year").agg(papers=("Title","count"), cites=("Cited by","sum")).reset_index()
by_year = by_year.sort_values("Year")

# Top autores
auth_flat = []
for a in df["Authors"].dropna():
    for auth in a.split(";"):
        auth_flat.append(auth.strip())
top_authors = Counter(auth_flat).most_common(10)

# Top países colaboradores
all_co = [c for cs in df["countries"] for c in cs if c != "Chile"]
top_countries = Counter(all_co).most_common(10)

# Top keywords
kw_flat = []
for k in df["Author Keywords"].dropna():
    for kw in k.split(";"):
        kw = kw.strip()
        if kw: kw_flat.append(kw.title())
top_kw = Counter(kw_flat).most_common(10)

# Tipos de doc
doc_types = df["Document Type"].value_counts().head(6)

# Papers más citados
top_cited = df.nlargest(10, "Cited by")[["Title","Authors","Year","Source title","Cited by"]].copy()
top_cited["Authors"] = top_cited["Authors"].apply(
    lambda x: x[:50] + "..." if isinstance(x, str) and len(x) > 50 else x
)
top_cited["Title"] = top_cited["Title"].apply(
    lambda x: x[:80] + "..." if isinstance(x, str) and len(x) > 80 else x
)

# Género
gender_counts = df["Sexo estimado primer autor"].value_counts() if "Sexo estimado primer autor" in df.columns else None

# Instituciones colaboradoras
def extract_institutions(aff):
    if not isinstance(aff, str): return []
    result = []
    for entry in aff.split(";"):
        parts = [p.strip() for p in entry.split(",") if len(p.strip()) > 4]
        if len(parts) >= 2:
            result.append(parts[-2] if parts[-1] in COUNTRIES else parts[-1])
    return result

inst_flat = [i for aff in df["Affiliations"].dropna() for i in extract_institutions(aff)]
inst_flat = [i for i in inst_flat if i not in COUNTRIES and len(i) > 5]
top_inst = Counter(inst_flat).most_common(10)

# ── workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1 — RESUMEN EJECUTIVO
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Resumen Ejecutivo"
ws1.sheet_view.showGridLines = False

for col, w in zip("ABCDEFGHIJKLMNO", [2,14,2,14,2,14,2,14,2,14,2,14,2,14,2]):
    ws1.column_dimensions[col].width = w
for r in range(1, 60):
    ws1.row_dimensions[r].height = 18

# Título principal
ws1.merge_cells("A1:O3")
t = ws1["A1"]
t.value = "ANÁLISIS BIBLIOMÉTRICO — UNIVERSIDAD DE CHILE"
t.font = Font(name="Calibri", bold=True, size=20, color=C_WHITE)
t.fill = fill(C_DARK)
t.alignment = align("center")

ws1.merge_cells("A4:O5")
s = ws1["A4"]
s.value = "Inteligencia Artificial · Período 2022–2025 · Fuente: Scopus"
s.font = Font(name="Calibri", size=12, color=C_WHITE, italic=True)
s.fill = fill(C_BLUE)
s.alignment = align("center")

# KPIs (row 7)
ws1.merge_cells("A6:O6"); ws1["A6"].fill = fill(C_WHITE)
ws1.row_dimensions[6].height = 10

kpi_box(ws1,  7,  2, total_docs,         "Publicaciones totales",    C_DARK)
kpi_box(ws1,  7,  6, total_cites,        "Citas totales",            C_BLUE)
kpi_box(ws1,  7, 10, h_index,            "Índice H",                 "0F766E")
kpi_box(ws1, 12,  2, avg_cites,          "Promedio citas / pub.",    "7C3AED")
kpi_box(ws1, 12,  6, f"{pct_intl}%",     "Colaboración internacional","B45309")
kpi_box(ws1, 12, 10, len(top_kw),        "Áreas temáticas (top)",    "BE123C")

ws1.row_dimensions[16].height = 10

# Tabla resumen indicadores
section_title(ws1, 17, 2, "Indicadores Clave por Año")
headers = ["Año","Publicaciones","Citas totales","Citas promedio"]
for c, h in enumerate(headers, 2):
    ws1.cell(18, c).value = h
style_header_row(ws1, 18, 5, C_BLUE)

for i, (_, row) in enumerate(by_year.iterrows(), 19):
    avg = round(row["cites"] / row["papers"], 1) if row["papers"] else 0
    vals = [int(row["Year"]), int(row["papers"]), int(row["cites"]), avg]
    for c, v in enumerate(vals, 2):
        ws1.cell(i, c).value = v
    style_data_row(ws1, i, 5, is_even=(i % 2 == 0))

# Fila totales
tr = 19 + len(by_year)
for c in range(2, 6):
    ws1.cell(tr, c).fill = fill(C_LBLUE)
    ws1.cell(tr, c).font = font(bold=True, size=10)
    ws1.cell(tr, c).alignment = align("center")
ws1.cell(tr, 2).value = "TOTAL"
ws1.cell(tr, 3).value = f"=SUM(C19:C{tr-1})"
ws1.cell(tr, 4).value = f"=SUM(D19:D{tr-1})"
ws1.cell(tr, 5).value = f"=ROUND(E{tr}/C{tr},1)" if False else round(avg_cites, 1)

# Nota metodológica
ws1.row_dimensions[tr + 2].height = 14
nota_row = tr + 2
ws1.merge_cells(f"B{nota_row}:N{nota_row+3}")
nota = ws1.cell(nota_row, 2)
nota.value = (
    "Nota metodológica: El corpus comprende 352 publicaciones recuperadas desde Scopus utilizando los términos "
    "'Artificial Intelligence', 'Machine Learning', 'Deep Learning', 'Generative AI' y 'ChatGPT', "
    "filtradas por afiliación a la Universidad de Chile (AF-ID 60072000) para el período 2022–2025. "
    "El sexo del primer autor fue estimado por inferencia a partir del nombre de pila (21,6% indeterminado)."
)
nota.font = Font(name="Calibri", size=9, color=C_DGRAY, italic=True)
nota.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
nota.fill = fill(C_GRAY)

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2 — PRODUCTIVIDAD
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Productividad")
ws2.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKLM", [2,22,12,2,18,12,2,18,12,2,2,2,2]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells("A1:M2")
t2 = ws2["A1"]
t2.value = "PRODUCTIVIDAD ACADÉMICA — U. Chile · IA 2022–2025"
t2.font = Font(name="Calibri", bold=True, size=16, color=C_WHITE)
t2.fill = fill(C_DARK)
t2.alignment = align("center")

# Tabla publicaciones por año (para chart)
section_title(ws2, 4, 2, "Publicaciones por Año")
ws2.cell(5, 2).value = "Año"
ws2.cell(5, 3).value = "Publicaciones"
style_header_row(ws2, 5, 2, C_BLUE)
ws2.cell(5, 2).alignment = align("center")
ws2.cell(5, 3).alignment = align("center")

for i, (_, row) in enumerate(by_year.iterrows(), 6):
    ws2.cell(i, 2).value = int(row["Year"])
    ws2.cell(i, 3).value = int(row["papers"])
    style_data_row(ws2, i, 3, is_even=(i % 2 == 0))

# Chart publicaciones por año
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Publicaciones por Año"
chart1.y_axis.title = "N° Publicaciones"
chart1.x_axis.title = "Año"
chart1.style = 10
chart1.width = 14; chart1.height = 10
data1 = Reference(ws2, min_col=3, min_row=5, max_row=5+len(by_year))
cats1 = Reference(ws2, min_col=2, min_row=6, max_row=5+len(by_year))
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.series[0].graphicalProperties.solidFill = C_BLUE
ws2.add_chart(chart1, "B10")

# Tabla top autores
section_title(ws2, 4, 5, "Top 10 Autores más Productivos")
ws2.cell(5, 5).value = "Autor"
ws2.cell(5, 6).value = "Publicaciones"
style_header_row(ws2, 5, 2, C_DARK)
ws2.cell(5, 5).alignment = align("center")
ws2.cell(5, 6).alignment = align("center")

for i, (auth, cnt) in enumerate(top_authors, 6):
    ws2.cell(i, 5).value = auth
    ws2.cell(i, 6).value = cnt
    style_data_row(ws2, i, 2, is_even=(i % 2 == 0))
    ws2.cell(i, 5).alignment = align("left")

# Chart top autores
chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Top 10 Autores"
chart2.y_axis.title = "Publicaciones"
chart2.style = 10
chart2.width = 14; chart2.height = 12
data2 = Reference(ws2, min_col=6, min_row=5, max_row=15)
cats2 = Reference(ws2, min_col=5, min_row=6, max_row=15)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.solidFill = "0F766E"
ws2.add_chart(chart2, "E10")

# Tabla tipos de documento
section_title(ws2, 4, 8, "Tipos de Documento")
ws2.cell(5, 8).value = "Tipo"
ws2.cell(5, 9).value = "N°"
style_header_row(ws2, 5, 2, C_DARK)
ws2.cell(5, 8).alignment = align("center")
ws2.cell(5, 9).alignment = align("center")

for i, (dtype, cnt) in enumerate(doc_types.items(), 6):
    ws2.cell(i, 8).value = dtype
    ws2.cell(i, 9).value = int(cnt)
    style_data_row(ws2, i, 2, is_even=(i % 2 == 0))
    ws2.cell(i, 8).alignment = align("left")

# Chart tipos (pie)
chart3 = PieChart()
chart3.title = "Tipos de Documento"
chart3.style = 10
chart3.width = 12; chart3.height = 10
data3 = Reference(ws2, min_col=9, min_row=5, max_row=5+len(doc_types))
cats3 = Reference(ws2, min_col=8, min_row=6, max_row=5+len(doc_types))
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
ws2.add_chart(chart3, "H10")

# Top keywords
section_title(ws2, 25, 2, "Top 10 Palabras Clave")
ws2.cell(26, 2).value = "Keyword"
ws2.cell(26, 3).value = "Frecuencia"
style_header_row(ws2, 26, 2, C_BLUE)
for i, (kw, cnt) in enumerate(top_kw, 27):
    ws2.cell(i, 2).value = kw
    ws2.cell(i, 3).value = cnt
    style_data_row(ws2, i, 3, is_even=(i % 2 == 0))
    ws2.cell(i, 2).alignment = align("left")

# Chart keywords
chart4 = BarChart()
chart4.type = "bar"
chart4.title = "Top 10 Keywords"
chart4.style = 10
chart4.width = 14; chart4.height = 10
data4 = Reference(ws2, min_col=3, min_row=26, max_row=36)
cats4 = Reference(ws2, min_col=2, min_row=27, max_row=36)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.series[0].graphicalProperties.solidFill = "7C3AED"
ws2.add_chart(chart4, "E25")

# Sexo estimado
if gender_counts is not None:
    section_title(ws2, 25, 8, "Sexo estimado primer autor")
    ws2.cell(26, 8).value = "Categoría"
    ws2.cell(26, 9).value = "N°"
    style_header_row(ws2, 26, 2, C_DARK)
    g_row = 27
    for cat, cnt in gender_counts.items():
        ws2.cell(g_row, 8).value = cat
        ws2.cell(g_row, 9).value = int(cnt)
        style_data_row(ws2, g_row, 2, is_even=(g_row % 2 == 0))
        ws2.cell(g_row, 8).alignment = align("left")
        g_row += 1
    chart5 = PieChart()
    chart5.title = "Sexo estimado primer autor"
    chart5.style = 10
    chart5.width = 12; chart5.height = 10
    data5 = Reference(ws2, min_col=9, min_row=26, max_row=26+len(gender_counts))
    cats5 = Reference(ws2, min_col=8, min_row=27, max_row=26+len(gender_counts))
    chart5.add_data(data5, titles_from_data=True)
    chart5.set_categories(cats5)
    ws2.add_chart(chart5, "H25")

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3 — IMPACTO
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Impacto")
ws3.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKLM", [2,22,14,12,2,18,14,2,14,14,2,2,2]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells("A1:M2")
t3 = ws3["A1"]
t3.value = "IMPACTO ACADÉMICO — U. Chile · IA 2022–2025"
t3.font = Font(name="Calibri", bold=True, size=16, color=C_WHITE)
t3.fill = fill(C_DARK)
t3.alignment = align("center")

# KPI impacto
kpi_box(ws3,  4, 2, total_cites,  "Citas totales",       C_BLUE)
kpi_box(ws3,  4, 6, h_index,      "Índice H",            "0F766E")
kpi_box(ws3,  4, 9, avg_cites,    "Promedio citas/pub.", "7C3AED")

# Citas por año (tabla)
section_title(ws3, 10, 2, "Citas por Año")
ws3.cell(11, 2).value = "Año"
ws3.cell(11, 3).value = "Publicaciones"
ws3.cell(11, 4).value = "Citas totales"
style_header_row(ws3, 11, 3, C_BLUE)

for i, (_, row) in enumerate(by_year.iterrows(), 12):
    ws3.cell(i, 2).value = int(row["Year"])
    ws3.cell(i, 3).value = int(row["papers"])
    ws3.cell(i, 4).value = int(row["cites"])
    style_data_row(ws3, i, 3, is_even=(i % 2 == 0))

# Chart línea citas por año
chart6 = LineChart()
chart6.title = "Evolución de Citas por Año"
chart6.y_axis.title = "Citas"
chart6.x_axis.title = "Año"
chart6.style = 10
chart6.width = 14; chart6.height = 10
data6 = Reference(ws3, min_col=4, min_row=11, max_row=11+len(by_year))
cats6 = Reference(ws3, min_col=2, min_row=12, max_row=11+len(by_year))
chart6.add_data(data6, titles_from_data=True)
chart6.set_categories(cats6)
chart6.series[0].graphicalProperties.line.solidFill = C_BLUE
chart6.series[0].graphicalProperties.line.width = 25000
ws3.add_chart(chart6, "B16")

# Tabla más citados
section_title(ws3, 10, 6, "Top 10 Artículos más Citados")
headers3 = ["Título","Autores","Año","Fuente","Citas"]
for c, h in enumerate(headers3, 6):
    ws3.cell(11, c).value = h
style_header_row(ws3, 11, 5, C_DARK)

for i, (_, row) in enumerate(top_cited.iterrows(), 12):
    vals = [row["Title"], row["Authors"], int(row["Year"]), row["Source title"], int(row["Cited by"])]
    for c, v in enumerate(vals, 6):
        ws3.cell(i, c).value = v
        ws3.cell(i, c).fill = fill(C_GRAY if i % 2 == 0 else C_WHITE)
        ws3.cell(i, c).font = font(size=9)
        ws3.cell(i, c).alignment = align("left", wrap=(c == 6))
        ws3.cell(i, c).border = border_bottom()
ws3.column_dimensions["F"].width = 40
ws3.column_dimensions["G"].width = 20
ws3.column_dimensions["I"].width = 22

# Tabla H-index explicación
section_title(ws3, 24, 2, "Cálculo del Índice H")
ws3.cell(25, 2).value = "Rank"
ws3.cell(25, 3).value = "Citas"
ws3.cell(25, 4).value = "¿H-paper?"
style_header_row(ws3, 25, 3, C_BLUE)
for i, (rank, cites) in enumerate(enumerate(cites_sorted[:15], 1), 26):
    ws3.cell(i, 2).value = rank
    ws3.cell(i, 3).value = int(cites)
    is_h = cites >= rank
    ws3.cell(i, 4).value = "✓" if is_h else ""
    bg = C_MINT if is_h else C_WHITE if i % 2 else C_GRAY
    for c in range(2, 5):
        ws3.cell(i, c).fill = fill(bg)
        ws3.cell(i, c).font = font(size=10, bold=is_h)
        ws3.cell(i, c).alignment = align("center")
        ws3.cell(i, c).border = border_bottom()

ws3.cell(41, 2).value = f"→ Índice H = {h_index}"
ws3.cell(41, 2).font = Font(name="Calibri", bold=True, size=12, color=C_DARK)

# ════════════════════════════════════════════════════════════════════════════
# HOJA 4 — COLABORACIÓN
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Colaboración")
ws4.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJKLM", [2,26,14,2,26,14,2,2,2,2,2,2,2]):
    ws4.column_dimensions[col].width = w

ws4.merge_cells("A1:M2")
t4 = ws4["A1"]
t4.value = "COLABORACIÓN ACADÉMICA — U. Chile · IA 2022–2025"
t4.font = Font(name="Calibri", bold=True, size=16, color=C_WHITE)
t4.fill = fill(C_DARK)
t4.alignment = align("center")

# KPI colaboración
kpi_box(ws4, 4, 2, len(intl_collabs), "Papers con colab. intl.", C_BLUE)
kpi_box(ws4, 4, 6, f"{pct_intl}%",    "% colaboración intl.",   "0F766E")
kpi_box(ws4, 4, 9, len(top_countries),"Países colaboradores",    "B45309")

# Tabla países
section_title(ws4, 10, 2, "Top 10 Países Colaboradores")
ws4.cell(11, 2).value = "País"
ws4.cell(11, 3).value = "N° Publicaciones"
style_header_row(ws4, 11, 2, C_BLUE)

for i, (co, cnt) in enumerate(top_countries, 12):
    ws4.cell(i, 2).value = co
    ws4.cell(i, 3).value = cnt
    style_data_row(ws4, i, 3, is_even=(i % 2 == 0))
    ws4.cell(i, 2).alignment = align("left")

# Chart países
chart7 = BarChart()
chart7.type = "bar"
chart7.title = "Países Colaboradores"
chart7.y_axis.title = "Publicaciones"
chart7.style = 10
chart7.width = 14; chart7.height = 12
data7 = Reference(ws4, min_col=3, min_row=11, max_row=21)
cats7 = Reference(ws4, min_col=2, min_row=12, max_row=21)
chart7.add_data(data7, titles_from_data=True)
chart7.set_categories(cats7)
chart7.series[0].graphicalProperties.solidFill = C_BLUE
ws4.add_chart(chart7, "B16")

# Tabla instituciones colaboradoras
section_title(ws4, 10, 5, "Top 10 Instituciones Asociadas")
ws4.cell(11, 5).value = "Institución"
ws4.cell(11, 6).value = "Menciones"
style_header_row(ws4, 11, 2, C_DARK)

for i, (inst, cnt) in enumerate(top_inst, 12):
    ws4.cell(i, 5).value = inst[:40]
    ws4.cell(i, 6).value = cnt
    style_data_row(ws4, i, 2, is_even=(i % 2 == 0))
    ws4.cell(i, 5).alignment = align("left")

# Chart instituciones
chart8 = BarChart()
chart8.type = "bar"
chart8.title = "Instituciones Colaboradoras"
chart8.style = 10
chart8.width = 14; chart8.height = 12
data8 = Reference(ws4, min_col=6, min_row=11, max_row=21)
cats8 = Reference(ws4, min_col=5, min_row=12, max_row=21)
chart8.add_data(data8, titles_from_data=True)
chart8.set_categories(cats8)
chart8.series[0].graphicalProperties.solidFill = "0F766E"
ws4.add_chart(chart8, "E16")

# Colab nacional vs internacional
section_title(ws4, 31, 2, "Colaboración Nacional vs Internacional")
ws4.cell(32, 2).value = "Tipo"
ws4.cell(32, 3).value = "Publicaciones"
style_header_row(ws4, 32, 2, C_BLUE)
only_chile = df[df["countries"].apply(lambda c: all(x == "Chile" for x in c) and len(c) > 0)]
solo = df[df["countries"].apply(lambda c: len(c) == 0)]
colabs = [
    ("Solo Chile (sin afil. intl.)", len(only_chile)),
    ("Colaboración internacional", len(intl_collabs)),
    ("Sin datos de país", len(solo)),
]
for i, (label, cnt) in enumerate(colabs, 33):
    ws4.cell(i, 2).value = label
    ws4.cell(i, 3).value = cnt
    style_data_row(ws4, i, 3, is_even=(i % 2 == 0))
    ws4.cell(i, 2).alignment = align("left")

chart9 = PieChart()
chart9.title = "Tipo de Colaboración"
chart9.style = 10
chart9.width = 12; chart9.height = 9
data9 = Reference(ws4, min_col=3, min_row=32, max_row=35)
cats9 = Reference(ws4, min_col=2, min_row=33, max_row=35)
chart9.add_data(data9, titles_from_data=True)
chart9.set_categories(cats9)
ws4.add_chart(chart9, "B36")

# ── guardar ───────────────────────────────────────────────────────────────────
wb.save(DEST)
print("Dashboard guardado:", DEST)
