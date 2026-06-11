"""
dashboard.py — Generador de Dashboard profesional en Excel tipo PowerBI
Crea reportes visuales con gráficos embebidos, tablas formateadas y diseño corporativo
"""

import io
from pathlib import Path
from datetime import datetime
import pandas as pd
import numpy as np

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import analysis as an
import plotly.graph_objects as go
import plotly.express as px

# ── COLORES CORPORATIVOS ──────────────────────────────────────────────────────
# Para openpyxl: sin # (ej: "1e3a5f")
THEME_PRIMARY = "1e3a5f"       # Azul oscuro
THEME_SECONDARY = "2563EB"     # Azul
THEME_ACCENT = "F59E0B"        # Naranja
THEME_SUCCESS = "10B981"       # Verde
THEME_LIGHT = "F8FAFC"         # Gris claro
THEME_TEXT = "1e293b"          # Texto oscuro

# Para Plotly: con # (ej: "#2563EB")
PLOTLY_PRIMARY = "#1e3a5f"
PLOTLY_SECONDARY = "#2563EB"
PLOTLY_ACCENT = "#F59E0B"
PLOTLY_SUCCESS = "#10B981"

PALETTE_COLORS = [PLOTLY_SECONDARY, PLOTLY_ACCENT, PLOTLY_SUCCESS, "#EF4444", "#8B5CF6", "#EC4899"]


class DashboardBuilder:
    """Constructor de dashboards profesionales en Excel."""

    def __init__(self, title="Scimap Report"):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self.title = title
        self.temp_images = []

    def add_sheet(self, name):
        """Crea nueva hoja."""
        return self.wb.create_sheet(name)

    def style_header(self, ws, title, subtitle="", row=1):
        """Encabezado profesional con fondo color tema."""
        # Título principal
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=THEME_PRIMARY, end_color=THEME_PRIMARY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 32

        if subtitle:
            ws.merge_cells(f"A{row+1}:H{row+1}")
            cell = ws[f"A{row+1}"]
            cell.value = subtitle
            cell.font = Font(name="Calibri", size=10, color="FFFFFF", italic=True)
            cell.fill = PatternFill(start_color=THEME_SECONDARY, end_color=THEME_SECONDARY, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row+1].height = 22
            return row + 2
        return row + 1

    def add_kpi_section(self, ws, kpis, row=1):
        """Tarjetas KPI en 4 columnas."""
        for idx, (label, value) in enumerate(kpis.items()):
            col = idx + 1

            # Valor
            ws.cell(row, col, value)
            cell = ws.cell(row, col)
            cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=THEME_SECONDARY, end_color=THEME_SECONDARY, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 28

            # Label
            ws.cell(row + 1, col, label)
            cell = ws.cell(row + 1, col)
            cell.font = Font(name="Calibri", size=9, color=THEME_TEXT)
            cell.fill = PatternFill(start_color=THEME_LIGHT, end_color=THEME_LIGHT, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row + 1].height = 20

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22

        return row + 2

    def add_chart(self, ws, fig, row, col_start=1, width=20, height=12):
        """Inserta gráfico Plotly como imagen PNG."""
        try:
            img_bytes = fig.to_image(format="png", width=1400, height=800)
            temp_path = Path(f"temp_chart_{len(self.temp_images)}.png")
            with open(temp_path, "wb") as f:
                f.write(img_bytes)

            img = XLImage(str(temp_path))
            img.width = width * 8
            img.height = height * 8

            col_letter = get_column_letter(col_start)
            ws.add_image(img, f"{col_letter}{row}")
            self.temp_images.append(temp_path)

            return row + int(height * 1.2)
        except Exception as e:
            print(f"Error: {e}")
            return row + 2

    def add_table(self, ws, df, start_row=1, start_col=1, max_rows=20, title=""):
        """Tabla formateada con estilo profesional."""
        if title:
            ws.merge_cells(f"A{start_row}:H{start_row}")
            cell = ws[f"A{start_row}"]
            cell.value = title
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=THEME_SECONDARY, end_color=THEME_SECONDARY, fill_type="solid")
            ws.row_dimensions[start_row].height = 20
            start_row += 1

        df_disp = df.head(max_rows)

        # Encabezados
        for col_idx, col_name in enumerate(df_disp.columns, 1):
            cell = ws.cell(start_row, col_idx, col_name)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=THEME_SECONDARY, end_color=THEME_SECONDARY, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[start_row].height = 18

        # Datos
        for row_idx, (_, row_data) in enumerate(df_disp.iterrows(), start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row_idx, col_idx, value)
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")

                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=THEME_LIGHT, end_color=THEME_LIGHT, fill_type="solid")

                # Borde sutil
                thin_border = Border(
                    left=Side(style="thin", color="E2E8F0"),
                    right=Side(style="thin", color="E2E8F0"),
                )
                cell.border = thin_border

        return start_row + len(df_disp) + 1

    def save(self, path):
        """Guarda el workbook."""
        self.wb.save(path)
        print(f"✓ Dashboard guardado: {path}")

        # Limpiar archivos temporales
        for temp_path in self.temp_images:
            try:
                temp_path.unlink()
            except:
                pass


def generate_dashboard(df, output_path="Scimap_Dashboard.xlsx"):
    """
    Genera dashboard profesional en Excel.

    df: DataFrame con datos bibliométricos
    """
    if df.empty:
        print("❌ DataFrame vacío")
        return None

    builder = DashboardBuilder("Scimap Report")

    # ── HOJA 1: RESUMEN EJECUTIVO ────────────────────────────────────────

    ws = builder.add_sheet("Resumen")
    row = builder.style_header(ws, "Scimap", f"Análisis Bibliométrico • {datetime.now().strftime('%d/%m/%Y')}")

    # KPIs
    h_idx = an.h_index(df)
    n_authors = len(df["authors"].explode().unique()) if "authors" in df.columns else 0
    total_cites = int(df["cited_by"].sum()) if "cited_by" in df.columns else 0

    row = builder.add_kpi_section(ws, {
        "Documentos": len(df),
        "Autores Únicos": n_authors,
        "Citas Totales": total_cites,
        "H-Index": int(h_idx),
    }, row=row)

    # Gráfico 1: Producción por año
    try:
        by_year = df.groupby("year").size().reset_index(name="count")
        fig = go.Figure(go.Bar(x=by_year["year"], y=by_year["count"], marker_color=PLOTLY_SECONDARY,
                               text=by_year["count"], textposition="outside"))
        fig.update_layout(title="Producción Científica por Año", showlegend=False, height=400,
                         plot_bgcolor="white", paper_bgcolor="white")
        row = builder.add_chart(ws, fig, row, width=18, height=10)
    except Exception as e:
        print(f"Error gráfico producción: {e}")

    # ── HOJA 2: AUTORES ──────────────────────────────────────────────────

    ws = builder.add_sheet("Autores")
    row = builder.style_header(ws, "Análisis de Autores")

    # Top 20 autores
    try:
        top_auth = an.top_authors(df).head(20)
        # Seleccionar solo columnas que existan
        cols = [c for c in ["author", "count", "citations"] if c in top_auth.columns]
        top_auth = top_auth[cols].reset_index(drop=True)
        top_auth.columns = [c.replace("author", "Autor").replace("count", "Documentos").replace("citations", "Citas") for c in top_auth.columns]
        row = builder.add_table(ws, top_auth, start_row=row, max_rows=20, title="Top 20 Autores Más Productivos")
    except Exception as e:
        print(f"Error tabla autores: {e}")

    # ── HOJA 3: KEYWORDS ─────────────────────────────────────────────────

    ws = builder.add_sheet("Keywords")
    row = builder.style_header(ws, "Palabras Clave")

    # Top keywords
    try:
        kw = an.top_keywords(df).head(20)
        fig = px.treemap(kw, path=["keyword"], values="count", color="count",
                        color_continuous_scale="Viridis", title="Top Keywords")
        fig.update_layout(height=500, paper_bgcolor="white")
        row = builder.add_chart(ws, fig, row, width=20, height=12)

        # Tabla complementaria
        kw_table = kw[["keyword", "count"]].reset_index(drop=True)
        kw_table.columns = ["Palabra Clave", "Frecuencia"]
        row = builder.add_table(ws, kw_table, start_row=row+1, max_rows=30)
    except Exception as e:
        print(f"Error keywords: {e}")

    # ── HOJA 4: GEOGRAFÍA ────────────────────────────────────────────────

    ws = builder.add_sheet("Países")
    row = builder.style_header(ws, "Distribución Geográfica")

    try:
        countries = df.explode("countries").groupby("countries").size().sort_values(ascending=False).head(25)
        countries_df = pd.DataFrame({"País": countries.index, "Documentos": countries.values}).reset_index(drop=True)

        # Gráfico de barras
        fig = px.bar(countries_df, x="Documentos", y="País", orientation="h",
                    color="Documentos", color_continuous_scale="Blues",
                    title="Top 25 Países por Producción")
        fig.update_layout(height=500, paper_bgcolor="white", plot_bgcolor="white")
        row = builder.add_chart(ws, fig, row, width=18, height=11)

        # Tabla
        row = builder.add_table(ws, countries_df, start_row=row, max_rows=25)
    except Exception as e:
        print(f"Error geografía: {e}")

    # ── HOJA 5: FUENTES ──────────────────────────────────────────────────

    ws = builder.add_sheet("Revistas")
    row = builder.style_header(ws, "Top Revistas/Fuentes")

    try:
        sources_df = an.top_sources(df).head(20)
        # Usar nombre de columna correcto
        col_name = sources_df.columns[0] if len(sources_df.columns) > 0 else "source"
        sources_df = sources_df.rename(columns={col_name: "source"})
        sources_df = sources_df[["source", "count"]].reset_index(drop=True) if "count" in sources_df.columns else sources_df
        sources_df.columns = ["Revista/Fuente", "Documentos"]

        fig = px.barh(sources_df, x="Documentos", y="Revista/Fuente",
                     color="Documentos", color_continuous_scale="Oranges",
                     title="Top 20 Revistas Científicas")
        fig.update_layout(height=500, paper_bgcolor="white", plot_bgcolor="white")
        row = builder.add_chart(ws, fig, row, width=18, height=11)

        row = builder.add_table(ws, sources_df, start_row=row, max_rows=20)
    except Exception as e:
        print(f"Error fuentes: {e}")

    # ── HOJA 6: DATOS COMPLETOS ──────────────────────────────────────────

    ws = builder.add_sheet("Datos")
    row = builder.style_header(ws, "Dataset Completo")

    cols_export = [c for c in ["title", "authors", "year", "journal", "doi", "cited_by", "keywords"]
                   if c in df.columns]

    try:
        df_export = df[cols_export].head(200).reset_index(drop=True)
        df_export.columns = [c.title() for c in df_export.columns]
        builder.add_table(ws, df_export, start_row=row, max_rows=200)
    except Exception as e:
        print(f"Error datos: {e}")

    builder.save(output_path)
    return output_path


if __name__ == "__main__":
    # Test
    import parser as bp
    from pathlib import Path

    sample_file = Path("data.bib")
    if sample_file.exists():
        df = bp.load_data(sample_file)
        generate_dashboard(df, "test_dashboard.xlsx")
