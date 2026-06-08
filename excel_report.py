"""
excel_report.py — Generador de reportes profesionales en Excel
Crea dashboards tipo PowerBI con gráficos, tablas y diseño visual
"""

import io
import base64
from pathlib import Path
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from pillow_heif import register_heif_opener

register_heif_opener()

# ── colores ───────────────────────────────────────────────────────────────────
COLORS = {
    "primary": "1e3a5f",      # Azul oscuro
    "secondary": "2563EB",    # Azul
    "accent": "F59E0B",       # Naranja/Amber
    "success": "10B981",      # Verde
    "warning": "EF4444",      # Rojo
    "light": "F8FAFC",        # Gris claro
    "text": "1e293b",          # Texto oscuro
}


class ExcelReportBuilder:
    """Generador de reportes Excel profesionales."""

    def __init__(self, title="Scimap Report", theme_color=COLORS["primary"]):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self.title = title
        self.theme_color = theme_color
        self.temp_images = []

    def add_sheet(self, name, index=None):
        """Agrega una nueva hoja al workbook."""
        if index is not None:
            ws = self.wb.create_sheet(name, index)
        else:
            ws = self.wb.create_sheet(name)
        return ws

    def set_column_widths(self, ws, widths: dict):
        """Configura anchos de columnas. {col_letter: width}"""
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def add_header(self, ws, title, subtitle="", row=1):
        """Agrega header profesional a una hoja."""
        # Título principal
        ws.merge_cells(f"A{row}:H{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = title
        title_cell.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=self.theme_color, end_color=self.theme_color, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 35

        if subtitle:
            ws.merge_cells(f"A{row+1}:H{row+1}")
            subtitle_cell = ws[f"A{row+1}"]
            subtitle_cell.value = subtitle
            subtitle_cell.font = Font(name="Calibri", size=11, color="64748B")
            subtitle_cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row+1].height = 20
            return row + 2

        return row + 1

    def add_kpi_cards(self, ws, kpis: dict, start_row=1, start_col=1):
        """
        Agrega tarjetas KPI en formato tabla.
        kpis: {"label": value, ...}
        """
        col = start_col
        for label, value in kpis.items():
            # Valor
            ws.cell(row=start_row, column=col, value=value)
            cell = ws.cell(row=start_row, column=col)
            cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLORS["secondary"], end_color=COLORS["secondary"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[start_row].height = 30

            # Etiqueta
            ws.cell(row=start_row+1, column=col, value=label)
            cell = ws.cell(row=start_row+1, column=col)
            cell.font = Font(name="Calibri", size=9, color="64748B")
            cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[start_row+1].height = 20

            col += 1

        return start_row + 2

    def add_chart_image(self, ws, fig, row, col, width=20, height=12):
        """
        Convierte un gráfico Plotly a imagen PNG e inserta en Excel.
        width y height en cm aprox.
        """
        try:
            # Convertir Plotly figure a PNG
            img_bytes = fig.to_image(format="png", width=1200, height=700)
            img_stream = io.BytesIO(img_bytes)
            img_stream.seek(0)

            # Guardar temporalmente
            temp_path = Path(f"/tmp/chart_{len(self.temp_images)}.png")
            with open(temp_path, "wb") as f:
                f.write(img_bytes)

            # Insertar en Excel
            img = XLImage(str(temp_path))
            img.width = width * 10  # Conversión aproximada a puntos
            img.height = height * 10

            col_letter = get_column_letter(col)
            ws.add_image(img, f"{col_letter}{row}")
            self.temp_images.append(temp_path)

            return row + (height // 2) + 2

        except Exception as e:
            print(f"Error insertando gráfico: {e}")
            return row + 2

    def add_data_table(self, ws, df, start_row=1, start_col=1, max_rows=20):
        """Inserta una tabla de datos con formato profesional."""
        df_display = df.head(max_rows)

        # Encabezados
        for col_idx, col_name in enumerate(df_display.columns, start=start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLORS["secondary"], end_color=COLORS["secondary"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos
        for row_idx, (_, row_data) in enumerate(df_display.iterrows(), start=start_row+1):
            for col_idx, value in enumerate(row_data, start=start_col):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Alternar colores de fila
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        ws.row_dimensions[start_row].height = 20

        return start_row + len(df_display) + 2

    def save(self, path):
        """Guarda el workbook a un archivo."""
        self.wb.save(path)
        print(f"Reporte guardado: {path}")

        # Limpiar imágenes temporales
        for temp_path in self.temp_images:
            try:
                temp_path.unlink()
            except:
                pass


def generate_report(df, output_path="scimap_report.xlsx"):
    """
    Genera un reporte Excel completo con dashboards profesionales.

    df: DataFrame con datos bibliométricos (contiene: title, authors, year,
        journal, doi, abstract, keywords, affiliations, countries, cited_by, etc.)
    """
    import analysis as an

    builder = ExcelReportBuilder(title="Scimap - Análisis Bibliométrico")

    # ── HOJA 1: RESUMEN EJECUTIVO ──────────────────────────────────────────

    ws = builder.add_sheet("Resumen Ejecutivo", index=0)
    row = builder.add_header(ws, "Scimap - Análisis Bibliométrico Inteligente",
                             f"Reporte generado: {datetime.now().strftime('%d/%m/%Y')}")

    # KPIs principales
    h_index = an.h_index(df)
    total_authors = len(df["authors"].explode().unique()) if "authors" in df.columns else 0
    total_citations = df["cited_by"].sum() if "cited_by" in df.columns else 0

    builder.add_kpi_cards(ws, {
        "Documentos": len(df),
        "Autores": total_authors,
        "Citas": int(total_citations),
        "H-Index": int(h_index),
    }, start_row=row, start_col=1)

    # Gráficos principales
    row = builder.add_header(ws, "", "Análisis Visual", row=row+3)

    # Producción por año
    try:
        fig_prod = _create_production_chart(df)
        row = builder.add_chart_image(ws, fig_prod, row, 1, width=18, height=10)
    except Exception as e:
        print(f"Error creando gráfico producción: {e}")

    # ── HOJA 2: AUTORES ───────────────────────────────────────────────────

    ws = builder.add_sheet("Autores", index=1)
    row = builder.add_header(ws, "Análisis de Autores")

    # Top autores tabla
    try:
        top_authors = an.top_authors(df).head(20)
        row = builder.add_data_table(ws, top_authors[["author", "count", "citations"]],
                                     start_row=row, max_rows=20)
    except:
        pass

    # ── HOJA 3: PALABRAS CLAVE ────────────────────────────────────────────

    ws = builder.add_sheet("Keywords", index=2)
    row = builder.add_header(ws, "Análisis de Palabras Clave")

    try:
        fig_kw = _create_keywords_chart(df)
        row = builder.add_chart_image(ws, fig_kw, row, 1, width=20, height=12)
    except Exception as e:
        print(f"Error creando gráfico keywords: {e}")

    # ── HOJA 4: GEOGRAFÍA ─────────────────────────────────────────────────

    ws = builder.add_sheet("Geografía", index=3)
    row = builder.add_header(ws, "Distribución Geográfica")

    try:
        top_countries = df.explode("countries").groupby("countries").size().sort_values(ascending=False).head(20)
        row = builder.add_data_table(ws, pd.DataFrame({
            "País": top_countries.index,
            "Documentos": top_countries.values
        }), start_row=row, max_rows=20)
    except:
        pass

    # ── HOJA 5: DATOS COMPLETOS ───────────────────────────────────────────

    ws = builder.add_sheet("Datos", index=4)
    row = builder.add_header(ws, "Dataset Completo")

    # Columnas a mostrar
    cols_show = ["title", "authors", "year", "journal", "cited_by", "keywords"]
    cols_display = [c for c in cols_show if c in df.columns]

    try:
        builder.add_data_table(ws, df[cols_display].head(100), start_row=row, max_rows=100)
    except:
        pass

    builder.save(output_path)
    return output_path


# Funciones helper para crear gráficos base
def _create_production_chart(df):
    """Gráfico de producción por año."""
    by_year = df.groupby("year").size()
    fig = go.Figure(go.Bar(x=by_year.index, y=by_year.values, marker_color=COLORS["secondary"]))
    fig.update_layout(title="Producción por Año", showlegend=False, height=400)
    return fig


def _create_authors_chart(df):
    """Top autores."""
    import analysis as an
    top = an.top_authors(df).head(15)
    fig = go.Figure(go.Barh(y=top["author"], x=top["count"], marker_color=COLORS["success"]))
    fig.update_layout(title="Top Autores", showlegend=False, height=400)
    return fig


def _create_keywords_chart(df):
    """Keywords treemap."""
    import analysis as an
    kw = an.top_keywords(df).head(20)
    fig = px.treemap(kw, path=["keyword"], values="count", color="count",
                     color_continuous_scale="Blues", title="Keywords")
    fig.update_layout(height=400)
    return fig


def _create_geography_chart(df):
    """Mapa de publicaciones por país."""
    countries_data = df.explode("countries").groupby("countries").size().reset_index(name="count")
    fig = go.Figure(data=go.Choropleth(
        locations=countries_data["countries"],
        z=countries_data["count"],
        colorscale="Blues"
    ))
    fig.update_layout(title="Publicaciones por País", height=400, geo=dict(scope="world"))
    return fig
